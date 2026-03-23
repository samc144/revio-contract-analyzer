"""
Revio Contract Review API - Production Version
Generates Excel contract risk analysis reports
"""

from flask import Flask, request, jsonify, send_file
import anthropic
import base64
import os
import json
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

app = Flask(__name__)

def create_excel_report(analysis_data, project_name, contract_type, filename):
    """Create Excel report with Risk Register, Executive Summary, and Key Contract Particulars"""
    
    wb = Workbook()
    wb.remove(wb.active)
    
    # Styles
    header_font = Font(bold=True, size=11, name='Arial')
    title_font = Font(bold=True, size=14, name='Arial')
    normal_font = Font(size=10, name='Arial')
    
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    critical_fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
    high_fill = PatternFill(start_color='FFC000', end_color='FFC000', fill_type='solid')
    medium_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
    low_fill = PatternFill(start_color='92D050', end_color='92D050', fill_type='solid')
    
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Risk Register sheet
    ws_risks = wb.create_sheet('Risk Register')
    
    headers = ['Risk ID', 'Category', 'Clause Reference', 'Clause Detail', 'Risk Explanation', 
               'Risk Rating (1-5)', 'Risk Level', 'Mitigation Required']
    
    for col_num, header in enumerate(headers, 1):
        cell = ws_risks.cell(row=1, column=col_num, value=header)
        cell.font = Font(bold=True, size=11, name='Arial', color='FFFFFF')
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = thin_border
    
    ws_risks.column_dimensions['A'].width = 10
    ws_risks.column_dimensions['B'].width = 15
    ws_risks.column_dimensions['C'].width = 25
    ws_risks.column_dimensions['D'].width = 40
    ws_risks.column_dimensions['E'].width = 50
    ws_risks.column_dimensions['F'].width = 12
    ws_risks.column_dimensions['G'].width = 15
    ws_risks.column_dimensions['H'].width = 50
    
    for idx, risk in enumerate(analysis_data.get('risks', []), 2):
        ws_risks.cell(row=idx, column=1, value=risk.get('risk_id', '')).font = normal_font
        ws_risks.cell(row=idx, column=2, value=risk.get('category', '')).font = normal_font
        ws_risks.cell(row=idx, column=3, value=risk.get('clause_ref', '')).font = normal_font
        ws_risks.cell(row=idx, column=4, value=risk.get('clause_detail', '')).font = normal_font
        ws_risks.cell(row=idx, column=5, value=risk.get('explanation', '')).font = normal_font
        ws_risks.cell(row=idx, column=6, value=risk.get('rating', '')).font = normal_font
        
        risk_level_cell = ws_risks.cell(row=idx, column=7, value=risk.get('risk_level', ''))
        risk_level_cell.font = Font(bold=True, size=10, name='Arial')
        
        risk_level = risk.get('risk_level', '').upper()
        rating = risk.get('rating', 0)
        if 'CRITICAL' in risk_level or rating == 5:
            risk_level_cell.fill = critical_fill
        elif 'HIGH' in risk_level or rating == 4:
            risk_level_cell.fill = high_fill
        elif 'MEDIUM' in risk_level or rating == 3:
            risk_level_cell.fill = medium_fill
        else:
            risk_level_cell.fill = low_fill
        
        ws_risks.cell(row=idx, column=8, value=risk.get('mitigation', '')).font = normal_font
        
        for col in [4, 5, 8]:
            ws_risks.cell(row=idx, column=col).alignment = Alignment(wrap_text=True, vertical='top')
    
    # Executive Summary sheet
    ws_summary = wb.create_sheet('Executive Summary', 0)
    
    row = 1
    title_cell = ws_summary.cell(row=row, column=1, value=f'COMMERCIAL RISK ANALYSIS - {project_name.upper()}')
    title_cell.font = title_font
    ws_summary.merge_cells(f'A{row}:F{row}')
    row += 2
    
    ws_summary.cell(row=row, column=1, value='Contract Type:').font = Font(bold=True, size=10, name='Arial')
    ws_summary.cell(row=row, column=2, value=contract_type).font = normal_font
    ws_summary.merge_cells(f'B{row}:F{row}')
    row += 1
    
    ws_summary.cell(row=row, column=1, value='Project:').font = Font(bold=True, size=10, name='Arial')
    ws_summary.cell(row=row, column=2, value=project_name).font = normal_font
    ws_summary.merge_cells(f'B{row}:F{row}')
    row += 1
    
    ws_summary.cell(row=row, column=1, value='Overall Risk Rating:').font = Font(bold=True, size=10, name='Arial')
    overall_rating = analysis_data.get('overall_rating', 'HIGH - Multiple Deal Breaker Issues')
    ws_summary.cell(row=row, column=2, value=overall_rating).font = Font(bold=True, size=10, name='Arial', color='FF0000')
    ws_summary.merge_cells(f'B{row}:F{row}')
    row += 2
    
    ws_summary.cell(row=row, column=1, value='RISK DISTRIBUTION BY SEVERITY').font = Font(bold=True, size=11, name='Arial')
    row += 1
    
    dist_headers = ['Risk Level', 'Count', 'Percentage', 'Examples']
    for col_num, header in enumerate(dist_headers, 1):
        cell = ws_summary.cell(row=row, column=col_num, value=header)
        cell.font = Font(bold=True, size=10, name='Arial', color='FFFFFF')
        cell.fill = header_fill
    row += 1
    
    for dist in analysis_data.get('risk_distribution', []):
        ws_summary.cell(row=row, column=1, value=dist.get('level', '')).font = normal_font
        ws_summary.cell(row=row, column=2, value=dist.get('count', 0)).font = normal_font
        ws_summary.cell(row=row, column=3, value=dist.get('percentage', '')).font = normal_font
        ws_summary.cell(row=row, column=4, value=dist.get('examples', '')).font = normal_font
        ws_summary.merge_cells(f'D{row}:F{row}')
        row += 1
    
    row += 1
    
    ws_summary.cell(row=row, column=1, value='CRITICAL ISSUES - DEAL BREAKERS (MUST NEGOTIATE BEFORE CONTRACT)').font = Font(bold=True, size=11, name='Arial')
    row += 1
    
    crit_headers = ['Risk ID', 'Issue', 'Clause', 'Required Action']
    for col_num, header in enumerate(crit_headers, 1):
        cell = ws_summary.cell(row=row, column=col_num, value=header)
        cell.font = Font(bold=True, size=10, name='Arial', color='FFFFFF')
        cell.fill = critical_fill
    row += 1
    
    for issue in analysis_data.get('critical_issues', []):
        ws_summary.cell(row=row, column=1, value=issue.get('risk_id', '')).font = normal_font
        ws_summary.cell(row=row, column=2, value=issue.get('issue', '')).font = normal_font
        ws_summary.cell(row=row, column=3, value=issue.get('clause', '')).font = normal_font
        ws_summary.cell(row=row, column=4, value=issue.get('action', '')).font = normal_font
        ws_summary.merge_cells(f'D{row}:F{row}')
        ws_summary.cell(row=row, column=4).alignment = Alignment(wrap_text=True, vertical='top')
        row += 1
    
    ws_summary.column_dimensions['A'].width = 20
    ws_summary.column_dimensions['B'].width = 25
    ws_summary.column_dimensions['C'].width = 25
    ws_summary.column_dimensions['D'].width = 50
    
    # Key Contract Particulars sheet
    ws_particulars = wb.create_sheet('Key Contract Particulars')
    
    row = 1
    ws_particulars.cell(row=row, column=1, value='KEY CONTRACT PARTICULARS (NON-RISK ITEMS)').font = Font(bold=True, size=12, name='Arial')
    ws_particulars.merge_cells(f'A{row}:E{row}')
    row += 1
    
    ws_particulars.cell(row=row, column=1, value='Important contractual items contractors must be aware of (not necessarily high risk)').font = Font(italic=True, size=9, name='Arial')
    ws_particulars.merge_cells(f'A{row}:E{row}')
    row += 1
    
    part_headers = ['Element', 'Detail', 'Contractor Action Required', 'Reference', 'Notes']
    for col_num, header in enumerate(part_headers, 1):
        cell = ws_particulars.cell(row=row, column=col_num, value=header)
        cell.font = Font(bold=True, size=10, name='Arial', color='FFFFFF')
        cell.fill = header_fill
    row += 1
    
    for part in analysis_data.get('key_particulars', []):
        ws_particulars.cell(row=row, column=1, value=part.get('element', '')).font = normal_font
        ws_particulars.cell(row=row, column=2, value=part.get('detail', '')).font = normal_font
        ws_particulars.cell(row=row, column=3, value=part.get('action', '')).font = normal_font
        ws_particulars.cell(row=row, column=4, value=part.get('reference', '')).font = normal_font
        ws_particulars.cell(row=row, column=5, value=part.get('notes', '')).font = normal_font
        
        for col in [2, 3, 5]:
            ws_particulars.cell(row=row, column=col).alignment = Alignment(wrap_text=True, vertical='top')
        row += 1
    
    ws_particulars.column_dimensions['A'].width = 25
    ws_particulars.column_dimensions['B'].width = 35
    ws_particulars.column_dimensions['C'].width = 40
    ws_particulars.column_dimensions['D'].width = 25
    ws_particulars.column_dimensions['E'].width = 35
    
    wb.save(filename)
    return filename

@app.route('/health', methods=['GET'])
def health():
    """Health check endpoint"""
    return jsonify({'status': 'healthy', 'service': 'Revio Contract Analyzer'}), 200

@app.route('/analyze', methods=['POST'])
def analyze():
    """API endpoint to analyze contracts"""
    try:
        api_key = request.form.get('api_key')
        project_name = request.form.get('project_name', 'Contract Review')
        contract_type = request.form.get('contract_type', 'Not Specified')
        payment_id = request.form.get('payment_id', 'N/A')
        file = request.files.get('file')
        
        if not api_key:
            return jsonify({'error': 'API key required'}), 400
        if not file:
            return jsonify({'error': 'Contract file required'}), 400
        
        # Read file once
        file_content = file.read()
        file.seek(0)
        
        # Determine media type - check content first, then filename
        # Check file magic bytes to determine actual type
        if file_content[:4] == b'%PDF':
            media_type = 'application/pdf'
        elif file_content[:2] == b'PK':  # DOCX files are ZIP format
            media_type = 'application/vnd.openxmlformats-officedocument.wordprocessingml.document'
        else:
            # Fallback to filename if we have it
            filename = file.filename.lower() if file.filename else ''
            if '.pdf' in filename or filename.endswith('.pdf'):
                media_type = 'application/pdf'
            elif '.docx' in filename or filename.endswith('.docx'):
                media_type = 'application/vnd.openxmlformats-officedocument.wordprocessingml.document'
            else:
                return jsonify({'error': 'Unsupported file type - only PDF and DOCX allowed'}), 400
        
        # Encode for API
        file_base64 = base64.b64encode(file_content).decode('utf-8')
        
        # Call Anthropic API
        client = anthropic.Anthropic(api_key=api_key)
        
        message = client.messages.create(
            model="claude-sonnet-4-20250514",
            max_tokens=8000,
            messages=[{
                "role": "user",
                "content": [
                    {"type": "document", "source": {"type": "base64", "media_type": media_type, "data": file_base64}},
                    {"type": "text", "text": """You are an expert quantity surveyor reviewing this contract for a UK contractor.

CRITICAL: Respond ONLY with valid JSON. No markdown, no code blocks, no extra text.

Provide comprehensive risk analysis in this JSON format:

{
  "overall_rating": "HIGH (4.1/5) - Brief description",
  "risk_distribution": [
    {"level": "CRITICAL (5)", "count": 6, "percentage": "16%", "examples": "Brief examples"},
    {"level": "HIGH (4)", "count": 11, "percentage": "29%", "examples": "Brief examples"},
    {"level": "MEDIUM-HIGH (3)", "count": 14, "percentage": "37%", "examples": "Brief examples"},
    {"level": "MEDIUM (2)", "count": 7, "percentage": "18%", "examples": "Brief examples"},
    {"level": "LOW (1)", "count": 0, "percentage": "0%", "examples": "None"}
  ],
  "critical_issues": [
    {"risk_id": "R001", "issue": "Short description", "clause": "Reference", "action": "Required action"}
  ],
  "risks": [
    {
      "risk_id": "R001",
      "category": "Liability/Payment/Termination/etc",
      "clause_ref": "Clause reference",
      "clause_detail": "Actual clause wording",
      "explanation": "Detailed risk explanation",
      "rating": 5,
      "risk_level": "CRITICAL",
      "mitigation": "Specific mitigation steps"
    }
  ],
  "key_particulars": [
    {
      "element": "Payment Terms/Working Hours/etc",
      "detail": "Specific details",
      "action": "What contractor must do",
      "reference": "Clause reference",
      "notes": "Additional notes"
    }
  ]
}

Requirements:
- Identify 20-40+ risks minimum
- Rate 1-5: 5=CRITICAL (deal-breaker), 4=HIGH, 3=MEDIUM-HIGH, 2=MEDIUM, 1=LOW
- Focus on: unlimited liability, payment terms, termination, warranties, insurance, indemnities
- Be specific about clauses
- Provide actionable mitigation
- Include key contract particulars

REMEMBER: ONLY JSON. No other text."""}
                ]
            }]
        )
        
        # Parse response
        response_text = message.content[0].text.strip()
        
        # Clean up response if it has code blocks
        if response_text.startswith('```'):
            lines = response_text.split('\n')
            response_text = '\n'.join(lines[1:-1]) if len(lines) > 2 else response_text
            if response_text.startswith('json'):
                response_text = response_text[4:].strip()
        
        analysis_data = json.loads(response_text)
        
        # Generate Excel file
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        safe_name = "".join(c for c in project_name if c.isalnum() or c in (' ', '-', '_')).rstrip()
        excel_filename = f"{payment_id}_{safe_name}_Risk_Analysis_{timestamp}.xlsx"
        excel_path = os.path.join('/tmp', excel_filename)
        
        create_excel_report(analysis_data, project_name, contract_type, excel_path)
        
        # Return file
        return send_file(
            excel_path,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=excel_filename
        )
        
    except anthropic.AuthenticationError:
        return jsonify({'error': 'Invalid API key'}), 401
    except anthropic.PermissionDeniedError:
        return jsonify({'error': 'No credits. Add funds to your account'}), 403
    except json.JSONDecodeError as e:
        return jsonify({'error': f'Failed to parse AI response: {str(e)}'}), 500
    except Exception as e:
        return jsonify({'error': str(e)}), 500

if __name__ == '__main__':
    port = int(os.environ.get('PORT', 5000))
    print("\n" + "="*60)
    print("🏗️  Revio Contract Risk Analyzer API")
    print("="*60)
    print(f"\nServer starting on port {port}...")
    print("\n📍 Server ready to receive requests")
    print("📊 Generates Excel reports with 3 sheets")
    print("="*60 + "\n")
    app.run(host='0.0.0.0', port=port, debug=False)
